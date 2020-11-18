from sklearn.feature_extraction.text import CountVectorizer
from sklearn.manifold import TSNE
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import nltk
from nltk.corpus import stopwords

wb = load_workbook('./homepage/test_input_file.xlsx')
sheet = wb['Лист1']

nltk.download("stopwords")
stopWords = stopwords.words('russian')

text_list_category1 = []
text_list_category2 = []
text_list_category3 = []

get_data1_from_xlsx = []
get_data2_from_xlsx = []
get_data3_from_xlsx = []


for i in range(1, sheet.max_row + 1):
	if sheet.cell(row=i, column=2).value == 'Категория 1':
		text_list_category1.append(sheet.cell(row=i, column=1).value)

for i in range(1, sheet.max_row + 1):
	if sheet.cell(row=i, column=2).value == 'Категория 2':
		text_list_category2.append(sheet.cell(row=i, column=1).value)

for i in range(1, sheet.max_row + 1):
	if sheet.cell(row=i, column=2).value == 'Категория 3':
		text_list_category3.append(sheet.cell(row=i, column=1).value)


def cat1_tsne(cat1 = text_list_category1):
	vect1 = CountVectorizer(encoding='koi8r', stop_words = stopWords).fit_transform(cat1)
	vect1_embedded = TSNE(n_components=2).fit_transform(vect1)
	return vect1_embedded.tolist()

def cat2_tsne(cat2 = text_list_category2):
	vect2 = CountVectorizer(encoding='koi8r', stop_words = stopWords).fit_transform(cat2)
	vect2_embedded = TSNE(n_components=2).fit_transform(vect2)
	return vect2_embedded.tolist()

def cat3_tsne(cat3 = text_list_category3):
	vect3 = CountVectorizer(encoding='koi8r', stop_words = stopWords).fit_transform(cat3)
	vect3_embedded = TSNE(n_components=2).fit_transform(vect3)
	return vect3_embedded.tolist()

def getData1():
	for i in range(len(text_list_category1)):
		coordinates = cat1_tsne()[i]
		text = text_list_category1[i]
		category = 'Категория 1'
		get_data1_from_xlsx.append(
			{
				'coordinates' : coordinates,
				'text' : text,
				'category' : category,
			})
	return get_data1_from_xlsx

def getData2():
	for i in range(len(text_list_category2)):
		coordinates = cat2_tsne()[i]
		text = text_list_category2[i]
		category = 'Категория 2'
		get_data2_from_xlsx.append(
			{
				'coordinates' : coordinates,
				'text' : text,
				'category' : category,
			})
	return get_data2_from_xlsx

def getData3():
	for i in range(len(text_list_category3)):
		coordinates = cat3_tsne()[i]
		text = text_list_category3[i]
		category = 'Категория 3'
		get_data3_from_xlsx.append(
			{
				'coordinates' : coordinates,
				'text' : text,
				'category' : category,
			})
	return get_data3_from_xlsx